#!/usr/bin/env python3
"""
Jira Tempo Importer - Import time entries from Google Sheets or local files to Tempo (Jira)
"""

import csv
import json
import os
import sys
import re
from datetime import datetime
from pathlib import Path
from typing import Optional

import requests

# Optional imports for Google Sheets
try:
    import gspread
    from google.oauth2.service_account import Credentials
    GSPREAD_AVAILABLE = True
except ImportError:
    GSPREAD_AVAILABLE = False

# Optional imports for Excel files
try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# Optional import for secure credential storage
try:
    import keyring
    KEYRING_AVAILABLE = True
except ImportError:
    KEYRING_AVAILABLE = False


# Config file path
CONFIG_FILE = Path(__file__).parent / "config.json"

# Data source types
SOURCE_GOOGLE_SHEETS = "google_sheets"
SOURCE_LOCAL_FILE = "local_file"

# Tempo API base URL
TEMPO_API_URL = "https://api.tempo.io/4"

# Keyring service name
KEYRING_SERVICE = "jira-tempo-importer"

# Logs directory
LOGS_DIR = Path(__file__).parent / "logs"

# Column indices (0-based) - adjust if your sheet has different structure
COL_DATE = 0        # A - datum (d.m. format, e.g., "1.12.")
COL_TASK_ID = 1     # B - task ID (Jira ticket, e.g., "PROJ-123")
COL_DESCRIPTION = 2 # C - description
COL_HOURS = 3       # D - pocet hodin (e.g., "2.5" or "2,5")
COL_IMPORTED = 4    # E - imported date

# Global config (loaded at runtime)
config: dict = {}

# Global logger for current import session
import_log: Optional[Path] = None


def init_import_log(dry_run: bool = False) -> Path:
    """Initialize a new import log file."""
    LOGS_DIR.mkdir(exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    prefix = "dry_run_" if dry_run else "import_"
    log_file = LOGS_DIR / f"{prefix}{timestamp}.log"

    with open(log_file, "w") as f:
        f.write(f"Jira Tempo Importer - {'Dry Run' if dry_run else 'Import'} Log\n")
        f.write(f"Started: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write("=" * 60 + "\n\n")

    return log_file


def log(message: str) -> None:
    """Write a message to the import log and print to console."""
    print(message)
    if import_log:
        with open(import_log, "a") as f:
            f.write(message + "\n")


def log_summary(imported: int, skipped: int, total_hours: float, dry_run: bool = False) -> None:
    """Write summary to the log file."""
    if not import_log:
        return

    with open(import_log, "a") as f:
        f.write("\n" + "=" * 60 + "\n")
        f.write(f"Finished: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write(f"Imported: {imported}\n")
        f.write(f"Skipped: {skipped}\n")
        f.write(f"Total hours: {total_hours:.2f}h\n")
        if dry_run:
            f.write("(Dry run - no actual changes made)\n")


def get_secret(key: str) -> Optional[str]:
    """Get a secret from keyring."""
    if not KEYRING_AVAILABLE:
        return None
    try:
        return keyring.get_password(KEYRING_SERVICE, key)
    except Exception:
        return None


def set_secret(key: str, value: str) -> bool:
    """Store a secret in keyring."""
    if not KEYRING_AVAILABLE:
        return False
    try:
        keyring.set_password(KEYRING_SERVICE, key, value)
        return True
    except Exception:
        return False


def delete_secret(key: str) -> None:
    """Delete a secret from keyring."""
    if not KEYRING_AVAILABLE:
        return
    try:
        keyring.delete_password(KEYRING_SERVICE, key)
    except Exception:
        pass


# Keys that should be stored securely in keyring
SECURE_KEYS = ["jira_api_token", "tempo_api_token"]


def load_config() -> dict:
    """Load configuration from config file and keyring."""
    cfg = {}
    if CONFIG_FILE.exists():
        with open(CONFIG_FILE, "r") as f:
            cfg = json.load(f)

    # Load secrets from keyring
    if KEYRING_AVAILABLE:
        for key in SECURE_KEYS:
            secret = get_secret(key)
            if secret:
                cfg[key] = secret

    return cfg


def save_config(cfg: dict) -> None:
    """Save configuration to config file and secrets to keyring."""
    # Separate secrets from regular config
    config_to_save = {}
    secrets_saved = []

    for key, value in cfg.items():
        if key in SECURE_KEYS:
            if KEYRING_AVAILABLE and set_secret(key, value):
                secrets_saved.append(key)
            else:
                # Fallback: save to config file if keyring unavailable
                config_to_save[key] = value
        else:
            config_to_save[key] = value

    with open(CONFIG_FILE, "w") as f:
        json.dump(config_to_save, f, indent=2)

    print(f"\nConfiguration saved to {CONFIG_FILE}")
    if secrets_saved:
        print(f"Secrets stored securely in system keyring: {', '.join(secrets_saved)}")


def test_jira_connection(base_url: str, email: str, api_token: str) -> bool:
    """Test Jira API connection."""
    url = f"{base_url}/rest/api/3/myself"
    try:
        response = requests.get(url, auth=(email, api_token), headers={"Accept": "application/json"})
        if response.status_code == 200:
            data = response.json()
            print(f"  Connected as: {data.get('displayName', 'Unknown')} ({data.get('emailAddress', '')})")
            return True
        else:
            print(f"  Failed: {response.status_code} - {response.text[:200]}")
            return False
    except Exception as e:
        print(f"  Error: {e}")
        return False


def test_tempo_connection(api_token: str) -> bool:
    """Test Tempo API connection."""
    url = f"{TEMPO_API_URL}/accounts"
    try:
        response = requests.get(url, headers={"Authorization": f"Bearer {api_token}"})
        if response.status_code == 200:
            print("  Connected to Tempo API")
            return True
        else:
            print(f"  Failed: {response.status_code} - {response.text[:200]}")
            return False
    except Exception as e:
        print(f"  Error: {e}")
        return False


def get_jira_account_id(base_url: str, email: str, api_token: str) -> Optional[str]:
    """Get Jira account ID for the authenticated user."""
    url = f"{base_url}/rest/api/3/myself"
    try:
        response = requests.get(url, auth=(email, api_token), headers={"Accept": "application/json"})
        if response.status_code == 200:
            return response.json().get("accountId")
    except Exception:
        pass
    return None


def interactive_setup() -> dict:
    """Run interactive setup to collect configuration."""
    print("\n" + "=" * 60)
    print("Jira Tempo Importer - Initial Setup")
    print("=" * 60)
    print("\nThis wizard will help you configure the importer.")

    cfg = {}

    # Jira configuration
    print("\n" + "-" * 40)
    print("JIRA CONFIGURATION")
    print("-" * 40)

    while True:
        cfg["jira_base_url"] = input("\nJira URL (e.g., https://company.atlassian.net): ").strip().rstrip("/")
        cfg["jira_email"] = input("Jira email: ").strip()
        cfg["jira_api_token"] = input("Jira API token (from https://id.atlassian.com/manage-profile/security/api-tokens): ").strip()

        print("\nTesting Jira connection...")
        if test_jira_connection(cfg["jira_base_url"], cfg["jira_email"], cfg["jira_api_token"]):
            # Get account ID automatically
            account_id = get_jira_account_id(cfg["jira_base_url"], cfg["jira_email"], cfg["jira_api_token"])
            if account_id:
                cfg["jira_account_id"] = account_id
                print(f"  Account ID: {account_id}")
            break
        else:
            retry = input("\nConnection failed. Retry? (y/n): ").strip().lower()
            if retry != "y":
                print("Setup cancelled.")
                sys.exit(1)

    # Tempo configuration
    print("\n" + "-" * 40)
    print("TEMPO CONFIGURATION")
    print("-" * 40)
    print("\nGet your Tempo API token from:")
    print("  Tempo > Settings > API Integration")

    while True:
        cfg["tempo_api_token"] = input("\nTempo API token: ").strip()

        print("\nTesting Tempo connection...")
        if test_tempo_connection(cfg["tempo_api_token"]):
            break
        else:
            retry = input("\nConnection failed. Retry? (y/n): ").strip().lower()
            if retry != "y":
                print("Setup cancelled.")
                sys.exit(1)

    # Data source configuration
    print("\n" + "-" * 40)
    print("DATA SOURCE CONFIGURATION")
    print("-" * 40)
    print("\nWhere do you want to load time entries from?")
    print("  1. Google Sheets")
    print("  2. Local CSV file")

    while True:
        choice = input("\nChoose [1/2]: ").strip()
        if choice == "1":
            if not GSPREAD_AVAILABLE:
                print("  Google Sheets support requires gspread and google-auth packages.")
                print("  Install with: pip install gspread google-auth")
                continue
            cfg["data_source"] = SOURCE_GOOGLE_SHEETS
            break
        elif choice == "2":
            cfg["data_source"] = SOURCE_LOCAL_FILE
            break
        else:
            print("  Invalid choice. Enter 1 or 2.")

    if cfg["data_source"] == SOURCE_GOOGLE_SHEETS:
        # Google Sheets configuration
        print("\n" + "-" * 40)
        print("GOOGLE SHEETS CONFIGURATION")
        print("-" * 40)

        while True:
            default_creds = "credentials.json"
            creds_input = input(f"\nGoogle credentials file path [{default_creds}]: ").strip()
            cfg["google_credentials_file"] = creds_input if creds_input else default_creds

            if os.path.exists(cfg["google_credentials_file"]):
                print(f"  Found: {cfg['google_credentials_file']}")
                break
            else:
                print(f"  File not found: {cfg['google_credentials_file']}")
                retry = input("Retry? (y/n): ").strip().lower()
                if retry != "y":
                    print("Setup cancelled.")
                    sys.exit(1)

        print("\nSpreadsheet ID can be found in the Google Sheets URL:")
        print("  https://docs.google.com/spreadsheets/d/SPREADSHEET_ID/edit")
        cfg["spreadsheet_id"] = input("\nSpreadsheet ID: ").strip()

    else:
        # Local file configuration
        print("\n" + "-" * 40)
        print("LOCAL FILE CONFIGURATION")
        print("-" * 40)
        print("\nSupported formats: CSV, XLSX, XLS")
        print("Expected columns: date, task_id, description, hours, imported")

        while True:
            cfg["local_file_path"] = input("\nPath to file: ").strip()

            if os.path.exists(cfg["local_file_path"]):
                ext = Path(cfg["local_file_path"]).suffix.lower()
                if ext not in (".csv", ".xlsx", ".xls", ".xlsm"):
                    print(f"  Unsupported format: {ext}")
                    print("  Supported: .csv, .xlsx, .xls")
                    continue
                if ext in (".xlsx", ".xls", ".xlsm") and not OPENPYXL_AVAILABLE:
                    print("  Excel support requires openpyxl.")
                    print("  Install with: pip install openpyxl")
                    continue
                print(f"  Found: {cfg['local_file_path']}")
                break
            else:
                print(f"  File not found: {cfg['local_file_path']}")
                retry = input("Retry? (y/n): ").strip().lower()
                if retry != "y":
                    print("Setup cancelled.")
                    sys.exit(1)

    print("\n" + "=" * 60)
    print("Setup complete!")
    print("=" * 60)

    return cfg


def ensure_config() -> dict:
    """Ensure configuration exists, run setup if needed."""
    cfg = load_config()

    required_keys = [
        "jira_base_url", "jira_email", "jira_api_token", "jira_account_id",
        "tempo_api_token", "data_source"
    ]

    # Add source-specific required keys
    if cfg.get("data_source") == SOURCE_GOOGLE_SHEETS:
        required_keys.extend(["google_credentials_file", "spreadsheet_id"])
    elif cfg.get("data_source") == SOURCE_LOCAL_FILE:
        required_keys.append("local_file_path")

    missing = [k for k in required_keys if not cfg.get(k)]

    if missing:
        print("Configuration incomplete or missing.")
        cfg = interactive_setup()
        save_config(cfg)

    return cfg


class LocalCSVWorksheet:
    """Wrapper for local CSV file to match gspread Worksheet interface."""

    def __init__(self, file_path: str):
        self.file_path = file_path
        self.title = Path(file_path).name
        self._data: list[list[str]] = []
        self._load()

    def _load(self) -> None:
        """Load CSV file into memory."""
        self._data = []
        with open(self.file_path, "r", encoding="utf-8") as f:
            # Try to detect delimiter
            sample = f.read(4096)
            f.seek(0)

            # Check for common delimiters
            if ";" in sample and "," not in sample:
                delimiter = ";"
            elif "\t" in sample and "," not in sample:
                delimiter = "\t"
            else:
                delimiter = ","

            reader = csv.reader(f, delimiter=delimiter)
            for row in reader:
                self._data.append(row)

    def get_all_values(self) -> list[list[str]]:
        """Return all values from the CSV."""
        return self._data

    def update_cell(self, row: int, col: int, value: str) -> None:
        """Update a cell and save the file."""
        # row is 1-based, col is 1-based
        row_idx = row - 1
        col_idx = col - 1

        # Extend rows if needed
        while len(self._data) <= row_idx:
            self._data.append([])

        # Extend columns if needed
        while len(self._data[row_idx]) <= col_idx:
            self._data[row_idx].append("")

        self._data[row_idx][col_idx] = value
        self._save()

    def _save(self) -> None:
        """Save data back to CSV file."""
        with open(self.file_path, "w", encoding="utf-8", newline="") as f:
            writer = csv.writer(f)
            writer.writerows(self._data)


class LocalExcelWorksheet:
    """Wrapper for local Excel file to match gspread Worksheet interface."""

    def __init__(self, file_path: str):
        if not OPENPYXL_AVAILABLE:
            raise RuntimeError("Excel support requires openpyxl. Install with: pip install openpyxl")

        self.file_path = file_path
        self._workbook = openpyxl.load_workbook(file_path)
        self._sheet = self._workbook.worksheets[0]  # Always use first worksheet
        self.title = self._sheet.title

    def get_all_values(self) -> list[list[str]]:
        """Return all values from the Excel sheet."""
        data = []
        for row in self._sheet.iter_rows():
            row_data = []
            for cell in row:
                value = cell.value
                if value is None:
                    row_data.append("")
                else:
                    row_data.append(str(value))
            data.append(row_data)
        return data

    def update_cell(self, row: int, col: int, value: str) -> None:
        """Update a cell and save the file."""
        self._sheet.cell(row=row, column=col, value=value)
        self._save()

    def _save(self) -> None:
        """Save workbook back to file."""
        self._workbook.save(self.file_path)


def get_local_worksheet(file_path: str):
    """Get appropriate worksheet wrapper based on file extension."""
    ext = Path(file_path).suffix.lower()

    if ext == ".csv":
        return LocalCSVWorksheet(file_path)
    elif ext in (".xlsx", ".xls", ".xlsm"):
        return LocalExcelWorksheet(file_path)
    else:
        raise ValueError(f"Unsupported file format: {ext}. Use .csv, .xlsx, or .xls")


def get_google_sheets_client():
    """Initialize Google Sheets client with service account credentials."""
    if not GSPREAD_AVAILABLE:
        raise RuntimeError("Google Sheets support not available. Install gspread and google-auth.")

    scopes = [
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive.readonly"
    ]

    credentials = Credentials.from_service_account_file(
        config["google_credentials_file"],
        scopes=scopes
    )

    return gspread.authorize(credentials)


def parse_date(date_str: str, year: Optional[int] = None) -> Optional[str]:
    """
    Parse date from 'd.m.' format to 'YYYY-MM-DD' format.
    If year is not provided, uses current year.
    """
    if not date_str or not date_str.strip():
        return None

    date_str = date_str.strip().rstrip(".")

    if year is None:
        year = datetime.now().year

    # Match patterns like "1.12" or "01.12"
    match = re.match(r"(\d{1,2})\.(\d{1,2})", date_str)
    if not match:
        return None

    day = int(match.group(1))
    month = int(match.group(2))

    try:
        date_obj = datetime(year, month, day)
        return date_obj.strftime("%Y-%m-%d")
    except ValueError:
        return None


# Cache for issue key to ID mapping
_issue_id_cache: dict[str, int] = {}


def parse_hours(hours_str: str) -> Optional[int]:
    """
    Parse hours string to seconds (Tempo API uses seconds).
    Handles both decimal point and comma as separator.
    """
    if not hours_str or not hours_str.strip():
        return None

    hours_str = hours_str.strip().replace(",", ".")

    try:
        hours = float(hours_str)
        return int(hours * 3600)  # Convert to seconds
    except ValueError:
        return None


def get_issue_id(issue_key: str) -> Optional[int]:
    """
    Get Jira issue ID from issue key using Jira REST API.
    Results are cached to avoid repeated API calls.

    Args:
        issue_key: Jira issue key (e.g., "PROJ-123")

    Returns:
        Numeric issue ID or None if not found
    """
    if issue_key in _issue_id_cache:
        return _issue_id_cache[issue_key]

    url = f"{config['jira_base_url']}/rest/api/3/issue/{issue_key}"

    auth = (config["jira_email"], config["jira_api_token"])
    headers = {
        "Accept": "application/json"
    }

    response = requests.get(url, auth=auth, headers=headers)

    if response.status_code == 200:
        data = response.json()
        issue_id = int(data["id"])
        _issue_id_cache[issue_key] = issue_id
        return issue_id
    else:
        print(f"    Warning: Could not find issue {issue_key} in Jira (status {response.status_code})")
        return None


def log_time_to_tempo(
    issue_key: str,
    date: str,
    time_spent_seconds: int,
    description: str
) -> dict:
    """
    Log time to Tempo via API.

    Args:
        issue_key: Jira issue key (e.g., "PROJ-123")
        date: Date in YYYY-MM-DD format
        time_spent_seconds: Time spent in seconds
        description: Work description

    Returns:
        API response as dict

    Raises:
        ValueError: If issue ID cannot be found
    """
    # Get numeric issue ID from Jira
    issue_id = get_issue_id(issue_key)
    if issue_id is None:
        raise ValueError(f"Could not find issue ID for {issue_key}")

    url = f"{TEMPO_API_URL}/worklogs"

    headers = {
        "Authorization": f"Bearer {config['tempo_api_token']}",
        "Content-Type": "application/json"
    }

    payload = {
        "issueId": issue_id,
        "timeSpentSeconds": time_spent_seconds,
        "startDate": date,
        "startTime": "09:00:00",  # Default start time
        "description": description,
        "authorAccountId": config["jira_account_id"]
    }

    response = requests.post(url, json=payload, headers=headers)
    response.raise_for_status()

    return response.json()


def process_worksheet(worksheet, dry_run: bool = False) -> tuple[int, int, float]:
    """
    Process worksheet and import unimported rows to Tempo.

    Args:
        worksheet: Google Sheets worksheet
        dry_run: If True, only show what would be imported without actually importing

    Returns:
        Tuple of (imported_count, skipped_count, total_hours)
    """
    # Get all values including headers
    all_values = worksheet.get_all_values()

    if len(all_values) <= 1:
        log("Worksheet is empty or has only headers.")
        return 0, 0, 0.0

    # Skip header row
    data_rows = all_values[1:]

    imported_count = 0
    skipped_count = 0
    total_hours = 0.0

    for row_idx, row in enumerate(data_rows, start=2):  # Start from 2 (1-based, skip header)
        # Ensure row has enough columns
        while len(row) < 5:
            row.append("")

        date_str = row[COL_DATE]
        task_id = row[COL_TASK_ID]
        description = row[COL_DESCRIPTION]
        hours_str = row[COL_HOURS]
        imported = row[COL_IMPORTED]

        # Parse hours early to count total hours for all rows
        time_seconds = parse_hours(hours_str)
        if time_seconds:
            total_hours += time_seconds / 3600

        # Skip if already imported
        if imported and imported.strip():
            skipped_count += 1
            continue

        # Skip empty rows
        if not date_str or not task_id or not hours_str:
            continue

        # Parse date
        parsed_date = parse_date(date_str)
        if not parsed_date:
            log(f"Row {row_idx}: Invalid date format '{date_str}', skipping.")
            skipped_count += 1
            continue

        # Validate hours (already parsed above)
        if not time_seconds:
            log(f"Row {row_idx}: Invalid hours format '{hours_str}', skipping.")
            skipped_count += 1
            continue

        # Clean task ID
        task_id = task_id.strip().upper()

        hours_display = time_seconds / 3600

        if dry_run:
            log(f"[DRY RUN] Row {row_idx}: Would import {task_id} - {parsed_date} - {hours_display}h - {description}")
            imported_count += 1
            continue

        try:
            log(f"Row {row_idx}: Importing {task_id} - {parsed_date} - {hours_display}h - {description[:30]}...")
            log_time_to_tempo(task_id, parsed_date, time_seconds, description)

            # Update the imported column with current date
            imported_date = datetime.now().strftime("%d.%m.%Y")
            worksheet.update_cell(row_idx, COL_IMPORTED + 1, imported_date)  # +1 for 1-based index

            log(f"  ✓ Successfully imported and marked as imported on {imported_date}")
            imported_count += 1

        except requests.exceptions.HTTPError as e:
            log(f"  ✗ Failed to import: {e}")
            if e.response is not None:
                log(f"    Response: {e.response.text}")
            skipped_count += 1
        except Exception as e:
            log(f"  ✗ Unexpected error: {e}")
            skipped_count += 1

    return imported_count, skipped_count, total_hours


def get_worksheet():
    """Get worksheet based on configured data source."""
    if config["data_source"] == SOURCE_GOOGLE_SHEETS:
        print("\nConnecting to Google Sheets...")
        client = get_google_sheets_client()

        print(f"Opening spreadsheet: {config['spreadsheet_id']}")
        spreadsheet = client.open_by_key(config["spreadsheet_id"])

        worksheets = spreadsheet.worksheets()
        worksheet = worksheets[0]
    else:
        print(f"\nLoading local file: {config['local_file_path']}")
        worksheet = get_local_worksheet(config["local_file_path"])

    print(f"Using: '{worksheet.title}'")
    return worksheet


def main():
    """Main entry point."""
    global config, import_log

    import argparse

    parser = argparse.ArgumentParser(description="Import time entries from Google Sheets or CSV to Tempo")
    parser.add_argument("--dry-run", action="store_true", help="Show what would be imported without actually importing")
    parser.add_argument("--year", type=int, help="Year for date parsing (default: current year)")
    parser.add_argument("--setup", action="store_true", help="Run setup wizard again")
    parser.add_argument("--file", type=str, help="Override: use this CSV file instead of configured source")
    args = parser.parse_args()

    print("=" * 60)
    print("Jira Tempo Importer")
    print("=" * 60)

    # Run setup if requested or config missing
    if args.setup:
        config = interactive_setup()
        save_config(config)
    else:
        config = ensure_config()

    # Override with --file argument if provided
    if args.file:
        if not os.path.exists(args.file):
            print(f"Error: File not found: {args.file}")
            sys.exit(1)
        ext = Path(args.file).suffix.lower()
        if ext not in (".csv", ".xlsx", ".xls", ".xlsm"):
            print(f"Error: Unsupported format: {ext}")
            print("Supported: .csv, .xlsx, .xls")
            sys.exit(1)
        if ext in (".xlsx", ".xls", ".xlsm") and not OPENPYXL_AVAILABLE:
            print("Error: Excel support requires openpyxl.")
            print("Install with: pip install openpyxl")
            sys.exit(1)
        config["data_source"] = SOURCE_LOCAL_FILE
        config["local_file_path"] = args.file
        print(f"\nUsing file override: {args.file}")

    # Initialize import log
    import_log = init_import_log(dry_run=args.dry_run)

    worksheet = get_worksheet()

    log("\nProcessing rows...")
    log("-" * 60)

    imported, skipped, total_hours = process_worksheet(worksheet, dry_run=args.dry_run)

    log("-" * 60)
    log(f"\nSummary:")
    log(f"  Imported:    {imported}")
    log(f"  Skipped:     {skipped}")
    log(f"  Total hours: {total_hours:.2f}h")

    if args.dry_run:
        log("\n(This was a dry run - no actual changes were made)")

    # Write final summary to log
    log_summary(imported, skipped, total_hours, dry_run=args.dry_run)
    print(f"\nLog saved to: {import_log}")


if __name__ == "__main__":
    main()
